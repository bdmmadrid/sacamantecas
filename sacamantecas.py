#! /usr/bin/env python3
"""
Saca las Mantecas.

This program reads an Excel file (xls/xlsx), containing a list of book titles,
each one with its signature and Manteca, which is an URI pointing to an entry
within some bibliographic catalogue where the book metadata can be obtained,
gets that metadata and adds it to each book, producing an output Excel file.

The Mantecas are processed according to profiles, which indicate how to properly
process the retrieved contents from the URIs, depending on the bibliographic
catalogue which is being processed. The proper profile is inferred from the URI
itself and resides in a separate file.

In short, it saca las Mantecas…

If the input file is not an Excel file, it is assumed it contains a list of
Mantecas, that is, a list of URIs pointing to bibliographic entries. In this
mode of operation the output file will not be an Excel file but another text
file containing the retrieved metadata for each entry. This is by design, so
profiles can be tested separately without the need to process and write Excel
files, or when the need arrives to process a new kind of URI in order to create
a new Manteca processing profile.
"""

# Current version…
__version__ = '1.0'

# Imports
import sys
import os.path
import errno
import logging
from logging.config import dictConfig
import traceback as tb
from shutil import copy2
from urllib.request import urlopen
from urllib.parse import urlparse, urlunparse
from urllib.error import URLError
import re
import time
from html.parser import HTMLParser
from msvcrt import getch
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import SheetTitleException, InvalidFileException
from openpyxl.utils.cell import get_column_letter
import win32con
import win32ui


# Class for the tag containing keys for the new metadata.
CONFIG_K_DIV_CLASS = 'auth'
# Class for the tag containing values for the new metadata.
CONFIG_V_DIV_CLASS = 'titn'
# Prefix to add to headers for the columns where the new metadata will go.
CONFIG_PREFIX = '[sm] '


# sys.modules[__name__].__file__ is used to determine the program's fully
# qualified directory and filename, so if it's not defined for some reason
# (which may happen...) it's better to break execution here.
try:
    PROGRAM_PATH = os.path.realpath(sys.modules['__main__'].__file__)
    PROGRAM_NAME = os.path.splitext(os.path.basename(PROGRAM_PATH))[0] + ' v' + __version__
except NameError:
    sys.exit('Error de inicialización del programa.')


# Check if platform is win32 or not.
if sys.platform != 'win32':
    sys.exit(f'{PROGRAM_NAME} solo funciona en la plataforma Win32.')


def error(message):
    """
    Show the error message 'message' on a Windows API MessageBox and stderr.

    This function is used when the end user needs to be signalled about a
    serious problem and the logging system is not direct enough or it has not
    been initialized yet.

    Since the logging system may not be initialized when calling this function,
    no logging functions should be used here.
    """
    print(f'\n\n*** Error en {PROGRAM_NAME}\n{message}', file=sys.stderr, end='')
    win32ui.MessageBox(message, f'Error en {PROGRAM_NAME}', win32con.MB_ICONERROR)


# Define the default exception hook.
def excepthook(exc_type, exc_value, exc_traceback):
    """Handle unhandled exceptions, default exception hook."""
    if isinstance(exc_value, OSError):
        # Handle OSError differently by giving more details.
        message = (
            f'Error inesperado del sistema operativo.\n'
            '['
            f'{exc_type.__name__}'
            f'{f"/{errno.errorcode[exc_value.errno]}" if exc_value.errno is not None else ""}'
            f'{f"/Win{exc_value.winerror}" if exc_value.winerror is not None else ""}'
            ']\n'
            f'{exc_value.strerror}.\n'
            f'{f"«{exc_value.filename}»" if exc_value.filename is not None else ""}'
            f'{f" -> «{exc_value.filename2}»" if exc_value.filename2 is not None else ""}'
        )
    else:
        message = f'Excepción {exc_type.__name__} sin gestionar.\n{exc_value}'

    message += '\n'
    message += '\n'.join([f'Línea {frame.lineno}: {frame.line}' for frame in tb.extract_tb(exc_traceback)]).rstrip()
    error(message)


# Install the default exception hook.
sys.excepthook = excepthook


def setup_logging(debugfile, logfile):
    """
    Sets up logging system, disabling all existing loggers.

    With the current configuration ALL logging messages are sent to 'debugfile',
    logging.INFO messages are sent to 'logfile' (timestamped), and the console
    (but not timestamped in this case).
    """
    logging_configuration = {
        'version': 1,
        'disable_existing_loggers': True,
        'formatters': {
            'debug': {
                'style': '{',
                'format': '{asctime}.{msecs:04.0f} [{levelname}] {message}',
                'datefmt': '%Y%m%d_%H%M%S'
            },
            'log': {
                'style': '{',
                'format': '{asctime} {message}',
                'datefmt': '%Y%m%d_%H%M%S'
            },
            'console': {
                'style': '{',
                'format': '{message}',
            },
        },
        'filters': {'info': {'()': lambda: lambda log_record: log_record.levelno == logging.INFO}},
        'handlers': {},
        'loggers': {
            '': {  # root logger.
                'level': 'NOTSET',
                'handlers': [],
                'propagate': False,
            },
        },
    }

    logging_configuration['handlers']['debugfile'] = {
        'level': 'NOTSET',
        'formatter': 'debug',
        'class': 'logging.FileHandler',
        'filename': debugfile,
        'mode': 'w',
        'encoding': 'utf8'
    }

    logging_configuration['handlers']['logfile'] = {
        'level': 'NOTSET',
        'formatter': 'log',
        'filters': ['info'],
        'class': 'logging.FileHandler',
        'filename': logfile,
        'mode': 'w',
        'encoding': 'utf8'
    }

    logging_configuration['handlers']['console'] = {
        'level': 'NOTSET',
        'formatter': 'console',
        'filters': ['info'],
        'class': 'logging.StreamHandler',
    }

    logging_configuration['loggers']['']['handlers'].append('debugfile')
    logging_configuration['loggers']['']['handlers'].append('logfile')
    logging_configuration['loggers']['']['handlers'].append('console')

    dictConfig(logging_configuration)
    logging.debug('Registro de depuración iniciado.')
    logging.debug('El registro de eventos se guardará en «%s».', logfile)
    logging.debug('El registro de depuración se guardará en «%s».', debugfile)
    logging.info(PROGRAM_NAME)
    for var in globals():
        if var.startswith('CONFIG_'):
            logging.debug('%s = %s', var, globals()[var])


class MantecaFile():
    """Abstract class to define an interface for Manteca files."""
    def __init__(self, filename):
        self.filename = filename

    def get_mantecas(self):
        """ Pure virtual function: get Mantecas from 'filename'."""
        raise NotImplementedError()

    def close(self):
        """ Pure virtual function: close 'filename'."""
        raise NotImplementedError()


class SkimmedFile():
    """Abstract class to define an interface for Manteca files."""
    def __init__(self, filename):
        self.filename = filename

    def add_metadata(self, row, uri, metadata):
        """Pure virtual function: add 'metadata' to Skimmed file."""
        raise NotImplementedError()

    def close(self):
        """ Pure virtual function: close 'filename'."""
        raise NotImplementedError()


class MantecaExcel(MantecaFile):
    """A class to represent Manteca Excel workbooks."""

    def __init__(self, *args, **kwargs):
        """Load the input Excel workbook."""
        super().__init__(*args, **kwargs)
        self.workbook = load_workbook(self.filename)
        # NOTE: not all sheets are processed, only the first one because it is
        # (allegedly) the one where the Manteca URIs for the items are.
        self.sheet = self.workbook.worksheets[0]
        logging.debug('La hoja con la que se trabajará es «%s»".', self.sheet.title)

    def get_mantecas(self):
        """
        Get the Mantecas found in the default worksheet. It's a generator.

        Returns a generator of (row, URI) tuples. Only the FIRST URI found in
        each row is considered and returned.
        """
        for row in self.sheet.rows:
            logging.debug('Procesando fila %s.', row[0].row)
            for cell in row:
                if cell.data_type != 's':
                    logging.debug('La celda %s no es de tipo cadena, será ignorada.', cell.coordinate)
                    continue
                if urlparse(cell.value).scheme.startswith('http'):
                    logging.debug('Se encontró un URI en la celda %s: %s', cell.coordinate, cell.value)
                    yield (cell.row, cell.value)
                    break

    def close(self):
        """Close the current workbook."""
        self.workbook.close()
        logging.debug('Fichero de Manteca cerrado.')


class SkimmedExcel(SkimmedFile):
    """A class to represent Skimmed (with 0% Manteca) Excel workbooks."""
    def __init__(self, *args, **kwargs):
        """Load the output Excel workbook."""
        super().__init__(*args, **kwargs)
        self.workbook = load_workbook(self.filename)
        # Keys are metadata names, values are the column where that metadata is stored.
        self.metadata_columns = {}
        # Style for cells on the header row.
        self.heading_style = {
            'font': Font(name='Calibri'),
            'fill': PatternFill(start_color='baddad', fill_type='solid'),
        }
        # NOTE: not all sheets are processed, only the first one because it is
        # (allegedly) the one where the Manteca URIs for the items are.
        self.sheet = self.workbook.worksheets[0]
        logging.debug('La hoja con la que se trabajará es «%s»".', self.sheet.title)

    def add_metadata(self, row, uri, metadata):
        """
        Add all specified 'metadata' to the default worksheet, at 'row'.

        The 'metadata' is a list of 'key'-'value' pairs.

        Each 'value' will be added in a new column if the 'key' doesn't already
        exists on the sheet, at the specified 'row'. The 'URI' is not used.

        Adds the header and styles it, also, if it doesn't exist.

        NOTE: the styling is just a best effort, and it's fragile. It depends on
        the sheet having headers on the first row, and the style used is that of
        the FIRST header.
        """
        logging.debug('Añadiendo metadatos para «%s».', uri)
        for key, value in metadata.items():
            key = CONFIG_PREFIX + key
            if key not in self.metadata_columns:
                logging.debug('Se encontró un metadato nuevo, «%s».', key)
                column = self.sheet.max_column + 1
                self.metadata_columns[key] = column
                logging.debug('El metadato «%s» irá en la columna %s.', key, get_column_letter(column))
                cell = self.sheet.cell(row=1, column=column, value=key)
                cell.font = self.heading_style['font']
                cell.fill = self.heading_style['fill']
                # Set column width.
                #
                # As per Excel specification, the width units are the width of
                # the zero character of the font used by the Normal style for a
                # workbook. So a column of width '10' would fit exactly 10 zero
                # characters in the font specified by the Normal style.
                #
                # No, no kidding.
                #
                # Since this width units are, IMHO, totally arbitrary, let's
                # choose an arbitrary column width. To wit, the Answer to the
                # Ultimate Question of Life, the Universe, and Everything.
                self.sheet.column_dimensions[get_column_letter(column)].width = 42
                # This is needed because sometimes Excel files are not properly
                # generated and the last column has a 'max' field too large, and
                # that has an unintended consequence: ANY change to the settings
                # of that column affects ALL the following ones whose index is
                # less than 'max'… So, it's better to fix that field.
                self.sheet.column_dimensions[get_column_letter(column)].max = column
            # Add the value to the proper column.
            logging.debug('Añadiendo metadato «%s» con valor «%s».', key, value)
            self.sheet.cell(row, self.metadata_columns[key], value=value)

    def close(self):
        """Close the current workbook, saving it."""
        self.workbook.save(self.filename)
        self.workbook.close()
        logging.debug('Workbook Excel guardado.')
        logging.debug('Fichero sin Manteca cerrado.')


class MantecaText(MantecaFile):
    """A class to represent Manteca text files."""
    def __init__(self, *args, **kwargs):
        """Load the input text file."""
        super().__init__(*args, **kwargs)
        self.file = open(self.filename, encoding='utf-8')  # pylint: disable=consider-using-with

    def get_mantecas(self):
        """
        Get the Mantecas found in the text file. It's a generator.

        Returns a generator of (row, URI) tuples, one per non empty file line.
        """
        for row, uri in enumerate(self.file.readlines(), start=1):
            uri = uri.strip()
            if uri:  # Do not return empty Mantecas.
                yield (row, uri)

    def close(self):
        self.file.close()
        logging.debug('Fichero de Manteca cerrado.')


class SkimmedText(SkimmedFile):
    """A class to represent Skimmed (with 0% Manteca) text files."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.file = open(self.filename, 'w', encoding='utf-8')  # pylint: disable=consider-using-with

    def add_metadata(self, row, uri, metadata):
        """
        Add all specified 'metadata' to this Skimmed text file.

        The 'metadata' is a list of 'key'-'value' pairs.

        The 'row' parameter is not used as a location where the 'data' will be
        added, since those are the file lines, and will be consecutive anyway.
        The 'row' parameter will be added at the beginning of each line as a
        reference only, followed by 'URI'. Then, the metadata will be more or
        less pretty-printed.
        """
        logging.debug('Añadiendo metadatos para «%s».', uri)
        self.file.write(f'[{row}] {uri}\n')
        for key, value in metadata.items():
            self.file.write(f'    {key}: {value}\n')

    def close(self):
        self.file.close()
        logging.debug('Fichero sin Manteca cerrado.')


class LibraryCatalogueHTMLParser(HTMLParser):
    """HTML parser for web pages of library catalogues."""

    def __init__(self, k_class, v_class, *args, **kwargs):
        """Initialize object."""
        super().__init__(*args, **kwargs)
        self.within_k_tag = False
        self.within_v_tag = False
        self.k_class = k_class
        self.v_class = v_class
        self.current_key = ''
        self.current_value = ''
        self.retrieved_metadata = {}

    def handle_starttag(self, tag, attrs):
        """Handle opening tags."""
        for attr in attrs:
            if attr[0] == 'class' and self.k_class in attr[1]:
                logging.debug('Se encontró una marca de clave.')
                self.within_k_tag = True
                self.current_key = ''
            if attr[0] == 'class' and self.v_class in attr[1]:
                logging.debug('Se encontró una marca de valor.')
                self.within_v_tag = True
                self.current_value = ''

    def handle_endtag(self, tag):
        """Handle closing tags."""
        if self.within_k_tag:
            self.within_k_tag = False
        if self.within_v_tag:
            self.within_v_tag = False
            # Metadata is only stored after getting the value.
            if not self.current_key:
                logging.debug('La clave estaba vacía.')
                self.current_key = '[vacío]'
            self.retrieved_metadata[self.current_key] = self.current_value
            self.current_key = ''
            self.current_value = ''

    def handle_data(self, data):
        """Handle data."""
        if self.within_k_tag or self.within_v_tag:
            # Clean up the received data by removing superfluous whitespace
            # characters, including newlines, carriage returns, etc.
            data = ' '.join(data.split())
            if not data:  # Ignore empty data
                return
        if self.within_k_tag:
            logging.debug('Se encontró la clave «%s».', data)
            self.current_key += data.rstrip(':')
        if self.within_v_tag:
            logging.debug('Se encontró el valor «%s».', data)
            if self.current_value:
                self.current_value += ' / '
            self.current_value += data

    def parse(self, contents):
        """Get some library item metadata from the 'contents' HTML."""
        self.retrieved_metadata.clear()
        self.feed(contents)
        self.close()
        return self.retrieved_metadata

    def error(self, _):
        """Override ParserBase abstract method."""


def retrieve_uri_contents(uri):
    """
    Retrieve contents from 'uri', performing redirections if needed.

    This function resolves meta-refresh redirection for 'uri', then gets the
    contents and decodes them using the detected charset, or utf-8 if none
    specified.

    NOTE about charset: if no charset is detected, then iso-8859-1 is used as
    default. Really, utf-8 should be a better default, because modern web pages
    may NOT specify any encoding if they are using utf-8 and it is identical to
    ascii in the 7-bit codepoints. The problem is that utf-8 will fail for pages
    encoded with iso-8859-1, and the vast majority of web pages processed will
    in fact use iso-8859-1 anyway.
    """
    try:
        with urlopen(uri) as request:
            # First, check if any redirection is needed and get the charset the easy way.
            logging.debug('Procesando URI «%s».', uri)
            contents = request.read()
            charset = request.headers.get_content_charset()
            match = re.search(rb'<meta http-equiv="refresh" content="[^;]+;\s*url=([^"]+)"', contents, re.I)
            if match:
                uri = urlparse(uri)
                uri = urlunparse((uri.scheme, uri.netloc, match.group(1).decode('ascii'), '', '', ''))
                logging.debug('Redirección -> «%s».', uri)
                with urlopen(uri) as redirected_request:
                    contents = redirected_request.read()
                    charset = redirected_request.headers.get_content_charset()
            else:
                logging.debug('El URI no está redirigido.')
    except ValueError as exc:
        if str(exc).startswith('unknown url type:'):
            raise URLError(f"El URI '{uri}' es de tipo desconocido.") from exc
        raise

    # In this point, we have the contents as a byte string.
    # If the charset is None, it has to be determined the hard way.
    if charset is None:
        # Next best thing, from the meta http-equiv="content-type".
        match = re.search(rb'<meta http-equiv="content-type".*charset=([^"]+)"', contents, re.I)
        if match:
            logging.debug('Charset detectado mediante meta http-equiv.')
            charset = match.group(1).decode('ascii')
        else:
            # Last resort, from some meta charset, if any…
            match = re.search(rb'<meta charset="([^"]+)"', contents, re.I)
            if match:
                logging.debug('Charset detectado mediante meta charset.')
                charset = match.group(1).decode('ascii')
            else:
                charset = 'latin_1'
                logging.debug('Usando charset por defecto.')
    else:
        logging.debug('Charset detectado en las cabeceras.')
    logging.debug('Contenidos codificados con charset «%s».', charset)

    # Decode the retrieved contents using the proper charset.
    return contents.decode(charset)


def main():  # pylint: disable=too-many-branches,too-many-statements
    """."""
    timestamp = time.strftime('%Y%m%d_%H%M%S')
    if len(sys.argv) < 2:
        # The input filename should be provided automatically if the program is
        # used as a drag'n'drop target, which is in fact the intended method of
        # operation.
        #
        # But the program can be also run by hand from a command prompt, so it
        # is better to give the end user a warning (well, error...) if the input
        # filename is missing.
        #
        # Also, at this point, the logging system is not initialized, so error
        # notification code has to be kept as simple as possible, using only
        # 'error()' to notify the end user of problems.
        error(
            'No se ha especificado un fichero de entrada para ser procesado.\n'
            '\n'
            'Arrastre y suelte un fichero de entrada sobre el icono del programa, '
            'o proporcione el nombre del fichero como argumento.'
        )
        return 1
    input_filename = sys.argv[1]
    output_filename = '_out'.join(os.path.splitext(input_filename))

    # Initialize logging system.
    # Generate the logging file names based upon input file name.
    debugfile = f'{os.path.splitext(input_filename)[0]}_debug_{timestamp}.txt'
    logfile = f'{os.path.splitext(input_filename)[0]}_log_{timestamp}.txt'
    setup_logging(debugfile, logfile)

    print()
    logging.info('El fichero de entrada es «%s».', input_filename)
    logging.info('El fichero de salida es «%s».', output_filename)

    error_message = ''
    try:
        if input_filename.endswith(('.xls', '.xlsx')):
            logging.debug('Los ficheros están en formato Excel.')
            mantecafile = MantecaExcel(input_filename)
            copy2(input_filename, output_filename)
            skimmedfile = SkimmedExcel(output_filename)
        else:
            logging.debug('Los ficheros están en formato texto.')
            mantecafile = MantecaText(input_filename)
            skimmedfile = SkimmedText(output_filename)

        print()
        logging.info('Sacando las mantecas:')
        parser = LibraryCatalogueHTMLParser(CONFIG_K_DIV_CLASS, CONFIG_V_DIV_CLASS)
        bad_uris = []
        for row, uri in mantecafile.get_mantecas():
            try:
                logging.info('  %s', uri)
                metadata = parser.parse(retrieve_uri_contents(uri))
                if not metadata:
                    bad_uris.append((uri, 'No se obtuvieron metadatos'))
                skimmedfile.add_metadata(row, uri, metadata)
            except ConnectionError:
                logging.debug('Error de conexión accediendo a «%s».', uri)
                bad_uris.append((uri, 'No se pudo conectar'))
            except URLError as exc:
                logging.debug('Error accediendo a «%s»: %s.', uri, exc.reason)
                bad_uris.append((uri, 'No se pudo acceder'))
        mantecafile.close()
        skimmedfile.close()
        if bad_uris:
            print()
            logging.info('Se encontraron problemas en los siguientes enlaces:')
        for uri, problem in bad_uris:
            logging.info('  [%s] %s.', uri, problem)
    except FileNotFoundError as exc:
        if exc.filename == input_filename:
            error_message = 'No se encontró el fichero de entrada.'
        else:  # This should not happen, so re-raise the exception.
            raise
    except PermissionError as exc:
        error_message = 'No hay permisos suficientes para '
        error_message += 'leer ' if exc.filename == input_filename else 'crear '
        error_message += 'el fichero de '
        error_message += 'entrada.' if exc.filename == input_filename else 'salida.'
    except (InvalidFileException, SheetTitleException):
        error_message = 'El fichero Excel de entrada es inválido.'
    except KeyboardInterrupt:
        logging.info('El usuario interrumpió la operación del programa.')
    finally:
        if error_message:
            logging.error('%s', error_message)
            error(error_message)
        print()
        logging.info('Proceso terminado.')
        logging.debug('Registro de eventos finalizado.')
        logging.shutdown()

    print('Pulse cualquier tecla para continuar...', end='', flush=True)
    getch()
    return 0 if not error_message else 1


if __name__ == '__main__':
    sys.exit(main())